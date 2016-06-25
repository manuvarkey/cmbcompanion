#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# bill_dialog.py
#  
#  Copyright 2014 Manu Varkey <manuvarkey@gmail.com>
#  
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#  
#  

import subprocess, threading, os, posixpath, platform, logging
import openpyxl

# Setup logger object
log = logging.getLogger(__name__)

## GLOBAL CONSTANTS

# Program name
PROGRAM_NAME = 'CMB Automiser'
# Item codes for data types
MEAS_NO = 1
MEAS_L = 2
MEAS_DESC = 3
MEAS_CUST = 4
# CMB error codes used for displaying info in main window
CMB_ERROR = -1
CMB_WARNING = -2
CMB_OK = 0
CMB_INFO = 0
# For tracking window management
CHILD_WINDOW = 1
PARENT_WINDOW = 0
main_hidden = 0
child_windows = 0
# Used in bill module for indicating type of bill
BILL_CUSTOM = 1
BILL_NORMAL = 2
# background colors for treeview
MEAS_COLOR_LOCKED = '#BABDB6'
MEAS_COLOR_NORMAL = '#FFFFFF'
MEAS_COLOR_SELECTED = '#729FCF'
# Timeout for killing Latex subprocess
LATEX_TIMEOUT = 300 # 5 minutes
# Item description wrap-width for screen purpose
CMB_DESCRIPTION_WIDTH = 60
CMB_DESCRIPTION_MAX_LENGTH = 1000
# Deviation statement
DEV_LIMIT_STATEMENT = 10
# List of units which will be considered as integer values
INT_ITEMS = ['point', 'points', 'pnt', 'pnts', 'number', 'numbers', 'no', 'nos', 'lot', 'lots',
             'lump', 'lumpsum', 'lump-sum', 'lump sum', 'ls', 'each','job','jobs','set','sets',
             'pair','pairs',
             'pnt.', 'no.', 'nos.', 'l.s.', 'l.s']
# String used for checking file version
PROJECT_FILE_VER = 'CMBAUTOMISER_FILE_REFERENCE_VER_3'
# Item codes for project global variables
global_vars = ['$cmbnameofwork$',
               '$cmbagency$',
               '$cmbagmntno$', 
               '$cmbsituation$',
               '$cmbdateofstart$',
               '$cmbdateofstartasperagmnt$',
               '$cmbissuedto$',
               '$cmbvarifyingauthority$',
               '$cmbvarifyingauthorityoffice$',
               '$cmbissuingauthority$',
               '$cmbissuingauthorityoffice$']
global_vars_captions = ['Name of Work', 
                        'Agency',
                        'Agreement Number',
                        'Situation',
                        'Date of Start',
                        'Date of start as per Agmnt.',
                        'CMB Issued to',
                        'Varifying Authority',
                        'Varifying Authority Office',
                        'Issuing Authority',
                        'Issuing Authority Office']
               
## GLOBAL VARIABLES

# Dict for storing saved settings
global_settings_dict = dict()

def set_global_platform_vars():
    """Setup global platform dependent variables"""
    
    if platform.system() == 'Linux':
        global_settings_dict['latex_path'] = 'pdflatex'
    elif platform.system() == 'Windows':
        global_settings_dict['latex_path'] = abs_path(
                    'miketex\\miktex\\bin\\pdflatex.exe')

## GLOBAL CLASSES

class Spreadsheet:
    """Manage input and output of spreadsheets"""
    
    def __init__(self, filename=None):
        if filename is not None:
            self.spreadsheet = openpyxl.load_workbook(filename)
        else:
            self.spreadsheet = openpyxl.Workbook()
        self.sheet = self.spreadsheet.active
    
    def save(self, filename):
        """Save worksheet to file"""
        self.spreadsheet.save(filename)
        
    # Sheet management
    
    def new_sheet(self):
        """Create a new sheet to spreadsheet and set as active"""
        self.sheet = self.spreadsheet.create_sheet()  
            
    def sheets(self):
        """Returns a list of sheetnames"""
        return self.spreadsheet.get_sheet_names()
        
    def length(self):
        """Get number of rows in sheet"""
        return len(self.sheet.rows)
        
    def set_title(self, title):
        """Set title of sheet"""
        self.sheet.title = title
        
    def set_column_widths(self, widths):
        """Set column widths of sheet"""
        for column, width in enumerate(widths, 1):
            col_letter = openpyxl.cell.get_column_letter(column)
            self.sheet.column_dimensions[col_letter].width = width
        
    def set_active_sheet(self, sheetref):
        """Set active sheet of spreadsheet"""
        sheetname = ''
        sheetno = None
        if type(sheetref) is int:
            sheetno = sheetref
        elif type(sheetref) is str:
            sheetname = sheetref
        
        if sheetname in self.sheets():
            self.sheet = self.spreadsheet[sheetname]
        elif sheetno is not None and sheetno < len(self.sheets()):
            self.sheet = self.spreadsheet[self.sheets()[sheetno]]
    
    def set_style(self, row, col, bold=False, wrap_text=True, horizontal='general'):
        """Set style of individual cell"""
        font = openpyxl.styles.Font(bold=bold)
        alignment = openpyxl.styles.Alignment(wrap_text=wrap_text, horizontal=horizontal)
        self.sheet.cell(row=row, column=col).font = font
        self.sheet.cell(row=row, column=col).alignment = alignment
        
    # Data addition functions
            
    def append(self, ss_obj):
        """Append an sheet to current sheet"""
        sheet = ss_obj.spreadsheet.active
        rowcount = self.length()
        for row_no, row in enumerate(sheet.rows, 1):
            for col_no, cell in enumerate(row, 1):
                self.sheet.cell(row=row_no+rowcount, column=col_no).value = cell.value
                self.sheet.cell(row=row_no+rowcount, column=col_no).style = cell.style
                
    def append_data(self, data, bold=False, wrap_text=True, horizontal='general'):
        """Append data to current sheet"""
        rowcount = self.length()
        self.insert_data(data, rowcount+1, 1, bold, wrap_text, horizontal)
    
    def insert_data(self, data, start_row=1, start_col=1, bold=False, wrap_text=True, horizontal='general'):
        """Insert data to current sheet"""
        # Setup styles
        font = openpyxl.styles.Font(bold=bold)
        alignment = openpyxl.styles.Alignment(wrap_text=wrap_text, horizontal=horizontal)
        # Apply data and styles
        for row_no, row in enumerate(data, start_row):
            for col_no, value in enumerate(row, start_col):
                self.sheet.cell(row=row_no, column=col_no).value = value
                self.sheet.cell(row=row_no, column=col_no).font = font
                self.sheet.cell(row=row_no, column=col_no).alignment = alignment
                
    def add_merged_cell(self, value, row=None, width=2, bold=False, wrap_text=True, horizontal='general'):
        """Add a merged cell of prescrbed width"""
        if row is None:
            rowstart = self.length() + 1
        else:
            rowstart = row
        self.sheet.merge_cells(start_row=rowstart,start_column=1,end_row=rowstart,end_column=width)
        self.__setitem__([rowstart,1], value)
        self.set_style(rowstart, 1, bold, wrap_text, horizontal)
    
    def __setitem__(self, index, value):
        """Set an individual cell"""
        self.sheet.cell(row=index[0], column=index[1]).value = value
        
    def __getitem__(self, index):
        """Set an individual cell"""
        return self.sheet.cell(row=index[0], column=index[1]).value
            
    # Bulk read functions
    
    def read_rows(self, columntypes = [], start=0, end=-1, left=0):
        """Read and validate selected rows from current sheet"""
        # Get count of rows
        rowcount = self.length()
        if end < 0 or end >= rowcount:
            count_actual = rowcount
        else:
            count_actual = end
        
        items = []
        for row in range(start, count_actual):
            cells = []
            skip = 0  # No of columns to be skiped ex. breakup, total etc...
            for columntype, i in zip(columntypes, list(range(left, len(columntypes)+left))):
                cell = self.sheet.cell(row = row + 1, column = i - skip + 1).value
                if columntype == MEAS_DESC:
                    if cell is None:
                        cell_formated = ""
                    else:
                        cell_formated = str(cell)
                elif columntype == MEAS_L:
                    if cell is None:
                        cell_formated = "0"
                    else:
                        try:  # try evaluating float
                            cell_formated = str(float(cell))
                        except:
                            cell_formated = '0'
                elif columntype == MEAS_NO:
                    if cell is None:
                        cell_formated = "0"
                    else:
                        try:  # try evaluating int
                            cell_formated = str(int(cell))
                        except:
                            cell_formated = '0'
                else:
                    cell_formated = ''
                    log.warning("Spreadsheet - Value skipped on import - " + str((row, i)))
                if columntype == MEAS_CUST:
                    skip = skip + 1
                cells.append(cell_formated)
            items.append(cells)
        return items


class LatexFile:
    """Class for formating and rendering of latex code"""
    
    def __init__(self, latex_buffer = ""):
        self.latex_buffer = latex_buffer
    
    # Inbuilt methods

    def clean_latex(self, text):
        """Replace special charchters with latex commands"""
        for splchar, replspelchar in zip(['\\', '#', '$', '%', '^', '&', '_', '{', '}', '~', '\n'],
                                         ['\\textbackslash ', '\# ', '\$ ', '\% ', '\\textasciicircum ', '\& ', '\_ ',
                                          '\{ ', '\} ', '\\textasciitilde ', '\\newline ']):
            text = text.replace(splchar, replspelchar)
        return text
        
    # Operator overloading
    
    def __add__(self,other):
        return LatexFile(self.latex_buffer + '\n' + other.latex_buffer)
        
    # Public members
    
    def get_buffer(self):
        """Get underlying latex buffer"""
        return self.latex_buffer
            
    def add_preffix_from_file(self,filename):
        """Add a latex file as preffix"""
        latex_file = open(filename,'r')
        self.latex_buffer = latex_file.read() + '\n' + self.latex_buffer
        latex_file.close()
        
    def add_suffix_from_file(self,filename):
        """Add a latex file as suffix"""
        latex_file = open(filename,'r')
        self.latex_buffer = self.latex_buffer + '\n' + latex_file.read()
        latex_file.close()
        
    def replace_and_clean(self, dic):
        """Replace items as per dictionary after cleaning special charachters"""
        for i, j in dic.items():
            j = self.clean_latex(j)
            self.latex_buffer = self.latex_buffer.replace(i, j)

    def replace(self, dic):
        """Replace items as per dictionary"""
        for i, j in dic.items():
            self.latex_buffer = self.latex_buffer.replace(i, j)
            
    def write(self, filename):
        """Write latex file to disk"""
        file_latex = open(filename,'w')
        file_latex.write(self.latex_buffer)
        file_latex.close()


class Command(object):
    """Runs a command in a seperate thread"""
    
    def __init__(self, cmd):
        """Initialises class with command to be executed"""
        self.cmd = cmd
        self.process = None

    def run(self, timeout):
        """Run set command with selected timeout"""
        def target():
            self.process = subprocess.Popen(self.cmd)
            log.info('Sub-process spawned - ' + str(self.process.pid))
            self.process.communicate()
        thread = threading.Thread(target=target)
        thread.start()

        thread.join(timeout)
        if thread.is_alive():
            log.error('Terminating sub-process exceeding timeout - ' + str(self.process.pid))
            self.process.terminate()
            thread.join()
            return -1
        return 0


## GLOBAL METHODS

def abs_path(*args):
    """Returns absolute path to the relative path provided"""
    return os.path.join(os.path.split(__file__)[0],*args)

def posix_path(*args):
    """Returns platform independent filename"""
    if platform.system() == 'Linux': 
        if len(args) > 1:
            return posixpath.join(*args)
        else:
            return args[0]
    elif platform.system() == 'Windows':
        if len(args) > 1:
            path = os.path.normpath(posixpath.join(*args))
        else:
            path = os.path.normpath(args[0])
        # remove any leading slash
        if path[0] == '\\':
            return path[1:]
        else:
            return path
            
def run_latex(folder, filename): 
    """Runs latex on file to folder in two passes"""
    if filename is not None:
        latex_exec = Command([global_settings_dict['latex_path'], '-interaction=batchmode', '-output-directory=' + folder, filename])
        # First Pass
        code = latex_exec.run(timeout=LATEX_TIMEOUT)
        if code == 0:
            # Second Pass
            code = latex_exec.run(timeout=LATEX_TIMEOUT)
            if code != 0:
                return CMB_ERROR
        else:
            return CMB_ERROR
    return CMB_OK

def clean_markup(text):
    """Clear markup text of special characters"""
    for splchar, replspelchar in zip(['&', '<', '>', ], ['&amp;', '&lt;', '&gt;']):
        text = text.replace(splchar, replspelchar)
    return text
    
